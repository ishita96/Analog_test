import os
from openpyxl import load_workbook

def getdirectory(path, extension="txt"):
    '''File fetch function'''
    count=0
    lensum=0
    print("-------------")
    for x in os.walk(path):#loop through each subdirectory
        for j in os.listdir(x[0]):#loop within a directory
            dir1=x[0]+"/"+j
            if j.endswith(extension):
                if extension=='xlsx':#special case of handling .xlsx file
                    count+=1
                    wb = load_workbook(dir1)
                    sheet = wb.worksheets[0]
                    row_count = sheet.max_row
                    lensum=lensum+row_count
                    print(j + " " +str(row_count))#row_count counts number of lines in xlsx
                else:
                    count+=1
                    with open(dir1,encoding='latin-1') as fp:
                        len1 = len(fp.readlines())
                        lensum=lensum+len1
                        print(j + " ", len1)#len1 counts number of lines in a text type file
    if count==0:
        print("No files found for the extension, try changing the extension to a valid one")#if it is not a valid input
    else:
        print("-----------------------")
        print("Number of files found :",count)
        print("Total number of lines :",lensum)
        print("Average lines per file :",lensum/count)

if __name__ == "__main__":
    path=input("Enter path:")
    extension=input("Extension: ")
    if path=="":
        print("Path is required")
    else:
        if extension:
            getdirectory(path,extension)
        else:
            getdirectory(path)